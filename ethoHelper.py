import openpyxl, os


# Python program that adds raw ethovision output into a checktime file and optionally creates awd files for each animal
def converter(inputFile, outputFile, startTime, date, awdFolder, buffer):
    actArr = []
    actArr.append([])
    animDict = {}

    print("Opening excel workbooks...")
    rdWB = openpyxl.load_workbook(inputFile)
    book = openpyxl.load_workbook(outputFile)
    newSheet = book.worksheets[0]
    worksheet = rdWB.worksheets[0]
    x = 4
    animNum = worksheet.cell(row=5, column=2).value
    animDict[animNum] = []

    print("Creating internal data storage...")
    correctCol = 4
    for col in range(4, 10):
        if isinstance(worksheet.cell(x + 1, col).value, float) or isinstance(worksheet.cell(x + 1, col).value, int):
            correctCol = col
    while worksheet.cell(row=(x + 1), column=1).value is not None:
        if animNum == worksheet.cell(row=x + 1, column=3).value:
            animDict[animNum].append(worksheet.cell(row=x + 1, column=correctCol).value)
        else:
            animNum = worksheet.cell(row=x + 1, column=3).value
            animDict[animNum] = []
            animDict[animNum].append(worksheet.cell(row=x + 1, column=correctCol).value)
        x += 1

    animNumSheet = 2
    startI = int(startTime[:2]) * 60 + int(startTime[3:]) + 2
    animList = animDict.keys()
    currCellVal = newSheet.cell(row=startI + 1, column=animNumSheet + 1).value
    while isinstance(currCellVal, float) or isinstance(currCellVal, int) or currCellVal == "-":
        currCellVal = newSheet.cell(row=startI + 1, column=animNumSheet + 1).value
        startI += 1
    currCount = 0

    print("Appending to checktime file...")
    for lst in animList:
        i = startI
        for count in range(len(animDict) + 2):
            if newSheet.cell(row=1, column=count + 1).value == lst:
                currCount = count
        for gapCount in range(int(buffer)):
            newSheet.cell(row=gapCount + i, column=currCount + 1).value = -1
        i = int(buffer) + i
        for data in animDict[lst]:
            newSheet.cell(row=i, column=currCount + 1).value = data
            i += 1

    print("Saving " + inputFile)
    rdWB.save(inputFile)
    print("Saving " + outputFile)
    book.save(outputFile)

    if date != "" and startTime != "":
        if not os.path.exists(awdFolder):
            os.mkdir(awdFolder)
        for lst in animList:
            path = awdFolder + "\\" + lst + ".awd"
            print("Generating awd file: " + path)
            if not os.path.exists(path):
                currFile = open(path, 'w')
                currFile.write(lst + '\n')
                currFile.write(date + '\n')
                currFile.write(startTime + '\n')
                currFile.write("4\n")
                currFile.write("0\n")
                currFile.write("Ignore\n")
                currFile.write("M\n")
            else:
                currFile = open(path, 'a')
            for gC in range(int(buffer)):
                currFile.write("-1\n")
            for data in animDict[lst]:
                currFile.write(str(data) + '\n')
            currFile.close()


def test():
    print("Test!")
    converter("ex.xlsx", "Checktime.xlsx", "19:00", "05/18/2018", "AWDFiles", 3)


def main():
    inputFile = ""
    outputFile = ""
    buffer = ""
    startTime = ""
    while inputFile == "":
        inputFile = input("Enter Ethovision output file (ex: rawInput): ")
    while outputFile == "":
        outputFile = input("Enter Checktime file (ex: checktime): ")
    while buffer == "":
        buffer = input("Please enter gap in minutes between today and yesterday(time that camera stopped; ex: 3):")
    while startTime == "":
        startTime = input("Please enter start time of first day(ex: 09:16,23:14): ")
    date = input("Please enter start date (Leave blank for no AWD file creation)(ex: 01/02/1970): ")
    if date != "":
        print("AWDs will be found and generated in format: <folder name>/<column name>.awd")
        awdFolder = input("Please enter folder name for AWD file storage(default: AWDFiles): ")
        if awdFolder == "":
            awdFolder = "AWDFiles"
    else:
        date = ""
        awdFolder = ""
    converter(inputFile + ".xlsx", outputFile + ".xlsx", startTime, date, awdFolder, buffer)


main()
